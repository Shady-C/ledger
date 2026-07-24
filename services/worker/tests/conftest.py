from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

TransactionSpec = tuple[date, str, Decimal, str | None, str]
ProcessedTransactionSpec = tuple[date, date, str, Decimal, str | None, str]


@pytest.fixture
def amex_workbook_bytes():
    def build(
        *,
        period_start: date,
        period_end: date,
        opening: Decimal,
        closing: Decimal,
        transactions: list[TransactionSpec],
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Statement"
        sheet.append(["American Express - Synthetic Test Statement"])
        sheet.append([f"Statement Period: {period_start.isoformat()} to {period_end.isoformat()}"])
        sheet.append(["Account", "••••1001"])
        sheet.append(["Opening Balance", f"CAD {opening}"])
        sheet.append(["Closing Balance", f"CAD {closing}"])
        sheet.append([])
        sheet.append([])
        sheet.append(["Date", "Description", "Amount", "Foreign Spend Amount", "Reference"])
        for booked, description, amount, foreign, reference in transactions:
            sheet.append([booked, description, str(amount), foreign, reference])
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    return build


@pytest.fixture
def amex_transaction_export_bytes():
    """Build a sanitized two-sheet workbook matching the real Amex export topology."""

    def build(
        *,
        period_start: date,
        period_end: date,
        opening: Decimal,
        closing: Decimal,
        transactions: list[ProcessedTransactionSpec],
        account_suffix: str = "54321",
    ) -> bytes:
        title = (
            "American Express Synthetic Card / "
            f"{period_start:%d %b %Y} to {period_end:%d %b %Y}"
        )
        workbook = Workbook()
        details = workbook.active
        details.title = "Transaction Details"
        details.append(["Transaction Details", title])
        details.append(["Prepared for", ""])
        details.append(["SYNTHETIC USER", ""])
        details.append(["Account Number", ""])
        details.append([f"XXXX-XXXXXX-{account_suffix}", ""])
        details.append([])
        details.append(
            [
                "Date",
                "Date Processed",
                "Description",
                "Amount",
                "Foreign Spend Amount",
                "Commission",
                "Exchange Rate",
                "Additional Information",
                "Merchant",
                "Address",
                "City / Province",
                "Postal Code",
                "Country",
                "Reference",
            ]
        )
        for booked, processed, description, amount, foreign, reference in transactions:
            details.append(
                [
                    booked,
                    processed,
                    description,
                    float(amount),
                    foreign or " ",
                    "",
                    "",
                    "",
                    description,
                    "",
                    "",
                    "",
                    "",
                    reference,
                ]
            )

        payments_and_credits = sum(
            (transaction[3] for transaction in transactions if transaction[3] < 0),
            Decimal("0"),
        )
        charges_and_adjustments = sum(
            (transaction[3] for transaction in transactions if transaction[3] > 0),
            Decimal("0"),
        )
        summary = workbook.create_sheet("Transaction Summary")
        summary.append(["Transaction Summary", title])
        summary.append(["Prepared for", ""])
        summary.append(["SYNTHETIC USER", ""])
        summary.append(["Account Number", ""])
        summary.append([f"XXXX-XXXXXX-{account_suffix}", ""])
        summary.append([])
        summary.append(["SUMMARY", ""])
        summary.append(["", "Total"])
        summary.append(["Last billed statement", float(opening)])
        summary.append(["Payments & Credits", float(payments_and_credits)])
        summary.append(["Charges & Adjustments", float(charges_and_adjustments)])
        summary.append(["Summary for this billed period", float(closing)])

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    return build
