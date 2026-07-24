from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from worker.adapters.amex_xlsx import AmexXlsxAdapter
from worker.adapters.base import AdapterError
from worker.models import Direction, ParsedFile
from worker.reconcile import reconcile_statement


def test_detects_buried_amex_header_and_preserves_card_signs(amex_workbook_bytes) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 1, 1),
        period_end=date(2026, 1, 31),
        opening=Decimal("100.00"),
        closing=Decimal("135.50"),
        transactions=[
            (date(2026, 1, 3), "Synthetic Market", Decimal("50.50"), None, "SYN-1"),
            (date(2026, 1, 8), "Payment Thank You", Decimal("-15.00"), None, "SYN-2"),
        ],
    )
    file = ParsedFile(name="synthetic.xlsx", content=content)
    adapter = AmexXlsxAdapter()

    assert adapter.detect(file) == pytest.approx(0.99)
    result = adapter.parse(file)

    assert result.statement.opening_balance == Decimal("100.00")
    assert result.statement.closing_balance == Decimal("135.50")
    assert [row.amount_native for row in result.rows] == [Decimal("50.50"), Decimal("-15.00")]
    assert [row.direction for row in result.rows] == [Direction.DEBIT, Direction.PAYMENT]


def test_parses_two_sheet_transaction_export_metadata_and_reconciles(
    amex_transaction_export_bytes,
) -> None:
    content = amex_transaction_export_bytes(
        period_start=date(2026, 6, 6),
        period_end=date(2026, 7, 5),
        opening=Decimal("2500.00"),
        closing=Decimal("2855.59"),
        transactions=[
            (
                date(2026, 6, 5),
                date(2026, 6, 6),
                "Synthetic Onboard Purchase",
                Decimal("35.67"),
                "USD 25.00",
                "SAFE-1",
            ),
            (
                date(2026, 7, 5),
                date(2026, 7, 5),
                "Synthetic Period Charge",
                Decimal("319.92"),
                None,
                "SAFE-2",
            ),
        ],
    )

    result = AmexXlsxAdapter().parse(
        ParsedFile(name="sanitized-transaction-export.xlsx", content=content)
    )
    reconciliation = reconcile_statement(result.statement, result.rows)

    assert result.statement.period_start == date(2026, 6, 6)
    assert result.statement.period_end == date(2026, 7, 5)
    assert result.statement.opening_balance == Decimal("2500.00")
    assert result.statement.closing_balance == Decimal("2855.59")
    assert result.statement.account_ref_masked == "••••54321"
    assert result.rows[0].booked_date == date(2026, 6, 5)
    assert result.rows[0].posted_date == date(2026, 6, 6)
    assert reconciliation.status == "ok"
    assert reconciliation.calculated_closing == Decimal("2855.59")


def test_accepts_extended_account_reference_on_same_row(
    amex_transaction_export_bytes,
) -> None:
    content = amex_transaction_export_bytes(
        period_start=date(2026, 6, 6),
        period_end=date(2026, 7, 5),
        opening=Decimal("0"),
        closing=Decimal("1"),
        transactions=[
            (
                date(2026, 7, 5),
                date(2026, 7, 5),
                "Synthetic Charge",
                Decimal("1"),
                None,
                "SAME-ROW-1",
            )
        ],
    )
    workbook = load_workbook(BytesIO(content))
    for sheet in workbook.worksheets:
        sheet["B4"] = sheet["A5"].value
        sheet["A5"] = None
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    result = AmexXlsxAdapter().parse(
        ParsedFile(name="same-row-account-reference.xlsx", content=output.getvalue())
    )

    assert result.statement.account_ref_masked == "••••54321"


@pytest.mark.parametrize(
    ("raw", "currency", "amount"),
    [
        ("USD 12.34", "USD", "12.34"),
        ("9.87 TZS", "TZS", "9.87"),
        ("US$5.25", "USD", "5.25"),
    ],
)
def test_parses_foreign_spend_enrichment(
    amex_workbook_bytes, raw: str, currency: str, amount: str
) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 2, 1),
        period_end=date(2026, 2, 28),
        opening=Decimal("0"),
        closing=Decimal("10"),
        transactions=[
            (date(2026, 2, 5), "Synthetic Hotel", Decimal("10"), raw, "FX-1"),
        ],
    )

    row = AmexXlsxAdapter().parse(ParsedFile(name="foreign.xlsx", content=content)).rows[0]

    assert row.currency_native == "CAD"
    assert row.enrichment["foreign_spend"] == {"amount": amount, "currency": currency}


def test_rejects_malformed_foreign_spend(amex_workbook_bytes) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 2, 1),
        period_end=date(2026, 2, 28),
        opening=Decimal("0"),
        closing=Decimal("10"),
        transactions=[(date(2026, 2, 5), "Hotel", Decimal("10"), "twelve", "FX-2")],
    )
    with pytest.raises(AdapterError, match="Foreign Spend"):
        AmexXlsxAdapter().parse(ParsedFile(name="bad.xlsx", content=content))


def test_rejects_transaction_or_balance_requiring_rounding(amex_workbook_bytes) -> None:
    bad_transaction = amex_workbook_bytes(
        period_start=date(2026, 3, 1),
        period_end=date(2026, 3, 31),
        opening=Decimal("0"),
        closing=Decimal("0.01"),
        transactions=[(date(2026, 3, 2), "Synthetic", Decimal("0.005"), None, "P-1")],
    )
    bad_balance = amex_workbook_bytes(
        period_start=date(2026, 3, 1),
        period_end=date(2026, 3, 31),
        opening=Decimal("0.005"),
        closing=Decimal("1.00"),
        transactions=[(date(2026, 3, 2), "Synthetic", Decimal("1.00"), None, "P-2")],
    )

    with pytest.raises(ValueError, match="exactly representable at two decimals"):
        AmexXlsxAdapter().parse(ParsedFile(name="bad-transaction.xlsx", content=bad_transaction))
    with pytest.raises(ValueError, match="exactly representable at two decimals"):
        AmexXlsxAdapter().parse(ParsedFile(name="bad-balance.xlsx", content=bad_balance))


def test_multiple_amex_alias_columns_fail_closed(amex_workbook_bytes) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        opening=Decimal("0"),
        closing=Decimal("1"),
        transactions=[(date(2026, 4, 2), "Synthetic", Decimal("1"), None, "A-1")],
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    header_row = next(
        row[0].row for row in sheet.iter_rows() if any(cell.value == "Date" for cell in row)
    )
    alias_column = sheet.max_column + 1
    sheet.cell(row=header_row, column=alias_column, value="Transaction Date")
    for row_index in range(header_row + 1, sheet.max_row + 1):
        sheet.cell(
            row=row_index,
            column=alias_column,
            value=sheet.cell(row=row_index, column=1).value,
        )
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(AdapterError, match="ambiguous booked date columns"):
        AmexXlsxAdapter().parse(ParsedFile(name="ambiguous-alias.xlsx", content=output.getvalue()))


def test_identical_description_and_merchant_columns_are_accepted(amex_workbook_bytes) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        opening=Decimal("0"),
        closing=Decimal("1"),
        transactions=[(date(2026, 4, 2), "Synthetic Market", Decimal("1"), None, "A-2")],
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    header_row = next(
        row[0].row for row in sheet.iter_rows() if any(cell.value == "Date" for cell in row)
    )
    merchant_column = sheet.max_column + 1
    sheet.cell(row=header_row, column=merchant_column, value="Merchant")
    sheet.cell(row=header_row + 1, column=merchant_column, value="  synthetic   MARKET ")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    result = AmexXlsxAdapter().parse(
        ParsedFile(name="equivalent-description.xlsx", content=output.getvalue())
    )

    assert [row.description_raw for row in result.rows] == ["Synthetic Market"]


def test_conflicting_description_and_merchant_columns_fail_closed(amex_workbook_bytes) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        opening=Decimal("0"),
        closing=Decimal("1"),
        transactions=[(date(2026, 4, 2), "Synthetic Market", Decimal("1"), None, "A-3")],
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    header_row = next(
        row[0].row for row in sheet.iter_rows() if any(cell.value == "Date" for cell in row)
    )
    merchant_column = sheet.max_column + 1
    sheet.cell(row=header_row, column=merchant_column, value="Merchant")
    sheet.cell(row=header_row + 1, column=merchant_column, value="Different Merchant")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(AdapterError, match="conflicting description columns"):
        AmexXlsxAdapter().parse(
            ParsedFile(name="conflicting-description.xlsx", content=output.getvalue())
        )


def test_amex_slash_order_is_inferred_across_period_and_all_rows(amex_workbook_bytes) -> None:
    content = amex_workbook_bytes(
        period_start=date(2026, 2, 1),
        period_end=date(2026, 4, 3),
        opening=Decimal("0"),
        closing=Decimal("3"),
        transactions=[
            (date(2026, 2, 13), "Synthetic One", Decimal("1"), None, "D-1"),
            (date(2026, 4, 3), "Synthetic Two", Decimal("2"), None, "D-2"),
        ],
    )
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    sheet.cell(row=2, column=1, value="Statement Period: 01/02/2026 to 03/04/2026")
    header_row = next(
        row[0].row for row in sheet.iter_rows() if any(cell.value == "Date" for cell in row)
    )
    sheet.cell(row=header_row + 1, column=1, value="13/02/2026")
    sheet.cell(row=header_row + 2, column=1, value="03/04/2026")
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    result = AmexXlsxAdapter().parse(ParsedFile(name="dmy-slash.xlsx", content=output.getvalue()))

    assert result.statement.period_start.isoformat() == "2026-02-01"
    assert result.statement.period_end.isoformat() == "2026-04-03"
    assert [row.booked_date.isoformat() for row in result.rows] == [
        "2026-02-13",
        "2026-04-03",
    ]
