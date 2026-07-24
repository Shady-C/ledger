from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from worker.models import AccountKind
from worker.pipeline import IngestionPipeline, JobRunner
from worker.repository import InMemoryRepository, Job
from worker.storage import MemoryObjectStore


def _golden_files(amex_workbook_bytes) -> dict[str, bytes]:
    return {
        "statements/synthetic-01.xlsx": amex_workbook_bytes(
            period_start=date(2026, 1, 1),
            period_end=date(2026, 1, 31),
            opening=Decimal("1000.00"),
            closing=Decimal("1200.00"),
            transactions=[
                (date(2026, 1, 3), "Synthetic Grocery Market", Decimal("500.00"), None, "G-1"),
                (date(2026, 1, 25), "Payment Thank You", Decimal("-300.00"), None, "G-2"),
            ],
        ),
        "statements/synthetic-02.xlsx": amex_workbook_bytes(
            period_start=date(2026, 2, 1),
            period_end=date(2026, 2, 28),
            opening=Decimal("1200.00"),
            closing=Decimal("1700.25"),
            transactions=[
                (
                    date(2026, 2, 8),
                    "Synthetic Hotel",
                    Decimal("750.50"),
                    "USD 520.00",
                    "G-3",
                ),
                (date(2026, 2, 20), "Merchant Refund", Decimal("-250.25"), None, "G-4"),
            ],
        ),
        "statements/synthetic-03.xlsx": amex_workbook_bytes(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            opening=Decimal("1700.25"),
            closing=Decimal("2855.59"),
            transactions=[
                (date(2026, 3, 9), "Synthetic Coffee Cafe", Decimal("1355.34"), None, "G-5"),
                (date(2026, 3, 27), "Autopay Payment", Decimal("-200.00"), None, "G-6"),
            ],
        ),
    }


def test_golden_three_statement_reconciliation_closes_at_2855_59(
    amex_workbook_bytes,
) -> None:
    objects = _golden_files(amex_workbook_bytes)
    repository = InMemoryRepository()
    pipeline = IngestionPipeline(store=MemoryObjectStore(objects), repository=repository)

    results = [
        pipeline.process_file(account_id="synthetic-account", file_key=file_key)
        for file_key in objects
    ]

    assert [result.reconcile["status"] for result in results] == ["ok", "ok", "ok"]
    assert results[-1].reconcile["reported_closing"] == "2855.59"
    assert results[-1].reconcile["calculated_closing"] == "2855.59"
    assert sum(result.added for result in results) == 6

    category_totals: dict[str, Decimal] = {}
    for row in repository.transactions.values():
        category_totals[row.category_name] = (
            category_totals.get(row.category_name, Decimal("0")) + row.amount_native
        )
    assert category_totals == {
        "Groceries": Decimal("500.00"),
        "Payments": Decimal("-500.00"),
        "Travel": Decimal("750.50"),
        "Refunds": Decimal("-250.25"),
        "Dining": Decimal("1355.34"),
    }
    merchant_totals: dict[str, Decimal] = {}
    for row in repository.transactions.values():
        merchant_totals[row.merchant_key] = (
            merchant_totals.get(row.merchant_key, Decimal("0")) + row.amount_native
        )
    assert merchant_totals == {
        "synthetic grocery market": Decimal("500.00"),
        "payment thank you": Decimal("-300.00"),
        "synthetic hotel": Decimal("750.50"),
        "merchant refund": Decimal("-250.25"),
        "synthetic coffee cafe": Decimal("1355.34"),
        "autopay payment": Decimal("-200.00"),
    }


def test_repeat_ingestion_adds_zero_rows(amex_workbook_bytes) -> None:
    objects = _golden_files(amex_workbook_bytes)
    key = next(iter(objects))
    repository = InMemoryRepository()
    pipeline = IngestionPipeline(store=MemoryObjectStore(objects), repository=repository)

    first = pipeline.process_file(account_id="synthetic-account", file_key=key)
    second = pipeline.process_file(account_id="synthetic-account", file_key=key)

    assert (first.added, first.skipped) == (2, 0)
    assert (second.added, second.skipped) == (0, 2)
    assert len(repository.transactions) == 2


def test_same_source_reimport_refreshes_summary_metadata_without_duplicate_transactions(
    amex_transaction_export_bytes,
) -> None:
    corrected = amex_transaction_export_bytes(
        period_start=date(2026, 6, 6),
        period_end=date(2026, 7, 5),
        opening=Decimal("2500.00"),
        closing=Decimal("2855.59"),
        transactions=[
            (
                date(2026, 6, 5),
                date(2026, 6, 6),
                "Synthetic Onboard Purchase",
                Decimal("35.67"),
                None,
                "REFRESH-1",
            ),
            (
                date(2026, 7, 5),
                date(2026, 7, 5),
                "Synthetic Period Charge",
                Decimal("319.92"),
                None,
                "REFRESH-2",
            ),
        ],
    )
    workbook = load_workbook(BytesIO(corrected))
    del workbook["Transaction Summary"]
    detail = workbook["Transaction Details"]
    processed_column = next(
        cell.column for cell in detail[7] if cell.value == "Date Processed"
    )
    for row_index in range(8, detail.max_row + 1):
        detail.cell(row=row_index, column=processed_column).value = None
    incomplete_output = BytesIO()
    workbook.save(incomplete_output)
    workbook.close()
    incomplete = incomplete_output.getvalue()

    key = "statements/same-source.xlsx"
    repository = InMemoryRepository(account_refs={"account": "••••0000"})
    first = IngestionPipeline(
        store=MemoryObjectStore({key: incomplete}),
        repository=repository,
    ).process_file(account_id="account", file_key=key)
    second = IngestionPipeline(
        store=MemoryObjectStore({key: corrected}),
        repository=repository,
    ).process_file(account_id="account", file_key=key)
    third = IngestionPipeline(
        store=MemoryObjectStore({key: incomplete}),
        repository=repository,
    ).process_file(account_id="account", file_key=key)

    statement = repository.statements[f"statement:account:{key}"]
    assert (first.added, first.reconcile["status"]) == (2, "pending")
    assert (second.added, second.skipped, second.reconcile["status"]) == (0, 2, "ok")
    assert (third.added, third.skipped, third.reconcile["status"]) == (0, 2, "ok")
    assert statement.metadata.opening_balance == Decimal("2500.00")
    assert statement.metadata.closing_balance == Decimal("2855.59")
    assert statement.status == "ok"
    assert len(repository.transactions) == 2
    assert {row.posted_date for row in repository.transactions.values()} == {
        date(2026, 6, 6),
        date(2026, 7, 5),
    }
    assert repository.account_refs["account"] == "••••54321"


def test_statement_account_reference_conflict_fails_closed(
    amex_transaction_export_bytes,
) -> None:
    content = amex_transaction_export_bytes(
        period_start=date(2026, 6, 6),
        period_end=date(2026, 7, 5),
        opening=Decimal("0"),
        closing=Decimal("1"),
        transactions=[
            (
                date(2026, 6, 6),
                date(2026, 6, 6),
                "Synthetic Charge",
                Decimal("1"),
                None,
                "ACCOUNT-1",
            )
        ],
    )
    for current_reference in ("••••99999", "••••94321"):
        repository = InMemoryRepository(account_refs={"account": current_reference})
        pipeline = IngestionPipeline(
            store=MemoryObjectStore({"statement.xlsx": content}),
            repository=repository,
        )

        with pytest.raises(ValueError, match="does not match selected account"):
            pipeline.process_file(account_id="account", file_key="statement.xlsx")

        assert repository.transactions == {}
        assert repository.statements == {}


def test_statement_account_reference_accepts_matching_last_four_digits(
    amex_transaction_export_bytes,
) -> None:
    content = amex_transaction_export_bytes(
        period_start=date(2026, 6, 6),
        period_end=date(2026, 7, 5),
        opening=Decimal("0"),
        closing=Decimal("1"),
        transactions=[
            (
                date(2026, 6, 6),
                date(2026, 6, 6),
                "Synthetic Charge",
                Decimal("1"),
                None,
                "ACCOUNT-2",
            )
        ],
    )
    repository = InMemoryRepository(account_refs={"account": "••••4321"})

    IngestionPipeline(
        store=MemoryObjectStore({"statement.xlsx": content}),
        repository=repository,
    ).process_file(account_id="account", file_key="statement.xlsx")

    assert repository.account_refs["account"] == "••••54321"


def test_job_runner_claims_processes_and_completes_an_ingest_job(
    amex_workbook_bytes,
) -> None:
    objects = _golden_files(amex_workbook_bytes)
    keys = list(objects)
    repository = InMemoryRepository(
        [
            Job(
                id="job-1",
                kind="ingest",
                payload={"account_id": "synthetic-account", "file_keys": keys},
            )
        ]
    )
    pipeline = IngestionPipeline(store=MemoryObjectStore(objects), repository=repository)
    runner = JobRunner(jobs=repository, pipeline=pipeline)

    assert runner.run_once() is True
    assert runner.run_once() is False
    assert repository.completed["job-1"]["status"] == "done"
    assert repository.completed["job-1"]["added"] == 6
    assert repository.completed["job-1"]["files"][2]["reconcile"]["status"] == "ok"
    assert repository.failed == {}


def test_job_runner_records_invalid_payload_as_failed() -> None:
    repository = InMemoryRepository([Job(id="job-2", kind="ingest", payload={})])
    pipeline = IngestionPipeline(store=MemoryObjectStore({}), repository=repository)

    assert JobRunner(jobs=repository, pipeline=pipeline).run_once() is True

    assert repository.failed["job-2"] == "invalid ingest job payload"


def test_gap_is_surfaced_and_cleared_when_missing_statement_arrives_out_of_order(
    amex_workbook_bytes,
) -> None:
    def statement(
        start: date,
        end: date,
        opening: str,
        closing: str,
        amount: str,
        reference: str,
    ) -> bytes:
        return amex_workbook_bytes(
            period_start=start,
            period_end=end,
            opening=Decimal(opening),
            closing=Decimal(closing),
            transactions=[
                (start, "Synthetic Grocery", Decimal(amount), None, reference),
            ],
        )

    objects = {
        "jan.xlsx": statement(date(2026, 1, 1), date(2026, 1, 31), "0", "10", "10", "JAN"),
        "feb.xlsx": statement(date(2026, 2, 1), date(2026, 2, 28), "10", "20", "10", "FEB"),
        "mar.xlsx": statement(date(2026, 3, 1), date(2026, 3, 31), "20", "30", "10", "MAR"),
    }
    repository = InMemoryRepository()
    pipeline = IngestionPipeline(store=MemoryObjectStore(objects), repository=repository)

    january = pipeline.process_file(account_id="account", file_key="jan.xlsx")
    march = pipeline.process_file(account_id="account", file_key="mar.xlsx")

    assert january.reconcile["status"] == "ok"
    assert march.reconcile["status"] == "gap"
    assert march.reconcile["coverage_gaps"] == [{"start": "2026-02-01", "end": "2026-02-28"}]
    assert repository.statements["statement:account:mar.xlsx"].status == "gap"

    february = pipeline.process_file(account_id="account", file_key="feb.xlsx")

    assert february.reconcile["status"] == "ok"
    assert february.reconcile["coverage_gaps"] == []
    assert {statement.status for statement in repository.statements.values()} == {"ok"}


def test_multi_file_job_preserves_successes_and_continues_after_safe_failure(
    amex_workbook_bytes,
    caplog,
) -> None:
    objects = _golden_files(amex_workbook_bytes)
    keys = list(objects)
    repository = InMemoryRepository(
        [
            Job(
                id="partial-job",
                kind="ingest",
                payload={
                    "account_id": "synthetic-account",
                    "file_keys": [keys[0], "missing-private-name.xlsx", keys[2]],
                },
            )
        ]
    )
    pipeline = IngestionPipeline(store=MemoryObjectStore(objects), repository=repository)

    with caplog.at_level(logging.WARNING, logger="worker.pipeline"):
        JobRunner(jobs=repository, pipeline=pipeline).run_once()

    result = repository.completed["partial-job"]
    assert result["status"] == "failed"
    assert result["added"] == 4
    assert [file["status"] for file in result["files"]] == ["done", "failed", "done"]
    assert result["files"][1]["reason"] == "source object is unavailable"
    assert "missing-private-name" not in repository.failed["partial-job"]
    assert len(repository.transactions) == 4
    assert repository.heartbeat_count >= 1
    failure_log = next(
        record for record in caplog.records if record.message == "file ingestion failed"
    )
    assert failure_log.exc_info is not None


def test_pipeline_uses_repository_account_kind_for_generic_signs() -> None:
    content = (
        b"Date,Description,Debit,Credit\n"
        b"2026-01-03,Synthetic withdrawal,25.00,\n"
        b"2026-01-04,Synthetic deposit,,15.00\n"
    )
    repository = InMemoryRepository(account_kinds={"asset-account": AccountKind.CHEQUING})
    pipeline = IngestionPipeline(
        store=MemoryObjectStore({"asset.csv": content}),
        repository=repository,
    )

    result = pipeline.process_file(account_id="asset-account", file_key="asset.csv")

    assert result.added == 2
    assert [row.amount_native for row in repository.transactions.values()] == [
        Decimal("-25.00"),
        Decimal("15.00"),
    ]
