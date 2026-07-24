"""Deterministic adapter for the American Express XLSX export."""

from __future__ import annotations

import re
from collections.abc import Sequence
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import polars as pl
from openpyxl import load_workbook

from worker.adapters.base import (
    AdapterError,
    extract_statement_metadata,
    infer_direction,
    infer_slash_date_order,
    locate_header_row,
    metadata_with_row_dates,
    normalize_header,
    parse_date,
    parse_decimal,
    resolve_unique_column,
    statement_period_date_values,
)
from worker.models import (
    AccountKind,
    ParsedFile,
    ParsedTransaction,
    ParseResult,
    StatementMetadata,
)

_DATE_HEADERS = ("date", "transaction date", "booked date")
_DESCRIPTION_HEADERS = ("description", "merchant", "details")
_AMOUNT_HEADERS = ("amount", "transaction amount", "amount cad")
_FOREIGN_HEADERS = ("foreign spend amount", "foreign amount")
_REFERENCE_HEADERS = ("reference", "reference id", "ref")
_POSTED_DATE_HEADERS = ("posted date", "posting date", "date processed")
_AMEX_TITLE_PERIOD = re.compile(
    r"\bAmerican\s+Express\b.*?/\s*"
    r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4})"
    r"\s+to\s+"
    r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4})",
    re.IGNORECASE,
)
_SUMMARY_OPENING_LABEL = "last billed statement"
_SUMMARY_CLOSING_LABEL = "summary for this billed period"

SheetRows = tuple[str, list[list[object]]]


class AmexXlsxAdapter:
    format = "xlsx"
    name = "amex_xlsx"

    def detect(self, file: ParsedFile) -> float:
        if file.extension not in {".xlsx", ".xlsm"}:
            return 0.0
        try:
            sheets = self._read_sheets(file)
            rows = _detail_rows(sheets)
        except Exception:
            return 0.0
        header = locate_header_row(
            rows,
            required_groups=(
                frozenset(_DATE_HEADERS),
                frozenset(_DESCRIPTION_HEADERS),
                frozenset(_AMOUNT_HEADERS),
            ),
        )
        if header is None:
            return 0.1
        normalized = {normalize_header(value) for value in rows[header]}
        amex_markers = {"foreign spend amount", "reference"}
        preamble = " ".join(str(cell) for row in rows[:header] for cell in row).lower()
        if "american express" in preamble or normalized.intersection(amex_markers):
            return 0.99
        return 0.72

    def parse(
        self,
        file: ParsedFile,
        *,
        account_kind: AccountKind = AccountKind.CREDIT_CARD,
    ) -> ParseResult:
        if account_kind is not AccountKind.CREDIT_CARD:
            raise AdapterError("Amex XLSX statements require a credit-card account")
        sheets = self._read_sheets(file)
        rows = _detail_rows(sheets)
        header_index = locate_header_row(
            rows,
            required_groups=(
                frozenset(_DATE_HEADERS),
                frozenset(_DESCRIPTION_HEADERS),
                frozenset(_AMOUNT_HEADERS),
            ),
        )
        if header_index is None:
            raise AdapterError("Amex XLSX header with Date, Description, and Amount was not found")

        headers = [normalize_header(cell) for cell in rows[header_index]]
        date_col = resolve_unique_column(headers, _DATE_HEADERS, field="booked date")
        assert date_col is not None
        description_col = _resolve_description_column(
            headers,
            rows=rows[header_index + 1 :],
            date_col=date_col,
        )
        amount_col = resolve_unique_column(headers, _AMOUNT_HEADERS, field="amount")
        foreign_col = resolve_unique_column(
            headers, _FOREIGN_HEADERS, field="foreign spend", required=False
        )
        reference_col = resolve_unique_column(
            headers, _REFERENCE_HEADERS, field="reference", required=False
        )
        posted_col = resolve_unique_column(
            headers, _POSTED_DATE_HEADERS, field="posted date", required=False
        )

        records: list[dict[str, Any]] = []
        for source_row in rows[header_index + 1 :]:
            if not any(value is not None and str(value).strip() for value in source_row):
                continue
            if _cell(source_row, date_col) in (None, ""):
                continue
            records.append(
                {
                    "booked": _cell(source_row, date_col),
                    "posted": _cell(source_row, posted_col),
                    "description": _cell(source_row, description_col),
                    "amount": _cell(source_row, amount_col),
                    "foreign": _cell(source_row, foreign_col),
                    "reference": _cell(source_row, reference_col),
                }
            )

        preamble = rows[: header_index + 1]
        slash_order = infer_slash_date_order(
            [
                value
                for record in records
                for value in (record.get("booked"), record.get("posted"))
                if value not in (None, "")
            ]
            + list(statement_period_date_values(preamble))
        )
        # Polars provides an explicit, columnar normalization boundary while
        # Decimal/date validation remains in the canonical Pydantic model.
        frame = pl.DataFrame(records, infer_schema_length=None) if records else pl.DataFrame()
        parsed_rows: list[ParsedTransaction] = []
        metadata = extract_statement_metadata(preamble, slash_date_order=slash_order)
        metadata = _merge_export_metadata(metadata, sheets)
        for record in frame.iter_rows(named=True):
            description = str(record["description"] or "").strip()
            if not description:
                raise AdapterError("transaction description is blank")
            amount = parse_decimal(record["amount"])
            enrichment: dict[str, Any] = {}
            if foreign := _parse_foreign_spend(record.get("foreign")):
                enrichment["foreign_spend"] = foreign
            parsed_rows.append(
                ParsedTransaction(
                    booked_date=parse_date(record["booked"], slash_order=slash_order),
                    posted_date=(
                        parse_date(record["posted"], slash_order=slash_order)
                        if record.get("posted")
                        else None
                    ),
                    description_raw=description,
                    amount_native=amount,
                    currency_native=metadata.currency,
                    external_ref=(
                        str(record["reference"]).strip() if record.get("reference") else None
                    ),
                    direction=infer_direction(description, amount),
                    enrichment=enrichment,
                )
            )

        if not parsed_rows:
            raise AdapterError("Amex XLSX contains no transaction rows")
        metadata = metadata_with_row_dates(metadata, [row.booked_date for row in parsed_rows])
        return ParseResult(adapter=self.name, rows=tuple(parsed_rows), statement=metadata)

    @staticmethod
    def _read_sheets(file: ParsedFile) -> list[SheetRows]:
        workbook = load_workbook(BytesIO(file.content), read_only=True, data_only=True)
        try:
            return [
                (sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)])
                for sheet in workbook.worksheets
            ]
        finally:
            workbook.close()


def _detail_rows(sheets: Sequence[SheetRows]) -> list[list[object]]:
    candidates: list[SheetRows] = []
    for title, rows in sheets:
        header = locate_header_row(
            rows,
            required_groups=(
                frozenset(_DATE_HEADERS),
                frozenset(_DESCRIPTION_HEADERS),
                frozenset(_AMOUNT_HEADERS),
            ),
        )
        if header is not None:
            candidates.append((title, rows))

    named = [rows for title, rows in candidates if normalize_header(title) == "transaction details"]
    if len(named) == 1:
        return named[0]
    if len(candidates) == 1:
        return candidates[0][1]
    if not candidates:
        raise AdapterError("Amex XLSX transaction-detail sheet was not found")
    raise AdapterError("Amex XLSX contains multiple possible transaction-detail sheets")


def _merge_export_metadata(
    metadata: StatementMetadata,
    sheets: Sequence[SheetRows],
) -> StatementMetadata:
    period = _export_period(sheets)
    account_ref = _export_account_ref(sheets)
    summary_rows = next(
        (rows for title, rows in sheets if normalize_header(title) == "transaction summary"),
        None,
    )
    opening = (
        _summary_amount(summary_rows, _SUMMARY_OPENING_LABEL) if summary_rows is not None else None
    )
    closing = (
        _summary_amount(summary_rows, _SUMMARY_CLOSING_LABEL) if summary_rows is not None else None
    )
    period_start = _consistent_value(
        "period start", metadata.period_start, period[0] if period else None
    )
    period_end = _consistent_value("period end", metadata.period_end, period[1] if period else None)
    return metadata.model_copy(
        update={
            "period_start": period_start,
            "period_end": period_end,
            "opening_balance": _consistent_value(
                "opening balance", metadata.opening_balance, opening
            ),
            "closing_balance": _consistent_value(
                "closing balance", metadata.closing_balance, closing
            ),
            "account_ref_masked": _consistent_account_reference(
                metadata.account_ref_masked, account_ref
            ),
        }
    )


def _export_period(sheets: Sequence[SheetRows]) -> tuple[date, date] | None:
    periods: set[tuple[date, date]] = set()
    for _title, rows in sheets:
        for row in rows:
            for value in row:
                if value is None or not (match := _AMEX_TITLE_PERIOD.search(str(value))):
                    continue
                periods.add((parse_date(match.group(1)), parse_date(match.group(2))))
    if len(periods) > 1:
        raise AdapterError("conflicting American Express export periods")
    return next(iter(periods), None)


def _export_account_ref(sheets: Sequence[SheetRows]) -> str | None:
    references: set[str] = set()
    for _title, rows in sheets:
        for row_index, row in enumerate(rows):
            label_indexes = [
                index
                for index, value in enumerate(row)
                if normalize_header(value) == "account number"
            ]
            for label_index in label_indexes:
                candidates = list(row[label_index + 1 :])
                if not any(value is not None and str(value).strip() for value in candidates):
                    candidates = _next_nonempty_row(rows, row_index + 1)
                for candidate in candidates:
                    if masked := _mask_account_ref(candidate):
                        references.add(masked)
    if len(references) > 1:
        raise AdapterError("conflicting American Express account references")
    return next(iter(references), None)


def _next_nonempty_row(rows: Sequence[Sequence[object]], start: int) -> list[object]:
    for row in rows[start:]:
        if any(value is not None and str(value).strip() for value in row):
            return list(row)
    return []


def _mask_account_ref(value: object) -> str | None:
    text = str(value or "").strip()
    match = re.search(r"(\d{4,6})\s*$", text)
    if match is None:
        return None
    suffix = match.group(1) if re.search(r"[xX*•]", text) else match.group(1)[-4:]
    return f"••••{suffix}"


def _summary_amount(rows: Sequence[Sequence[object]], label: str) -> Decimal | None:
    values: set[Decimal] = set()
    for row in rows:
        for index, cell in enumerate(row):
            if normalize_header(cell) != label:
                continue
            for candidate in row[index + 1 :]:
                if candidate is None or not str(candidate).strip():
                    continue
                values.add(parse_decimal(candidate))
                break
    if len(values) > 1:
        raise AdapterError(f"conflicting American Express {label} values")
    return next(iter(values), None)


def _consistent_value[T](field: str, existing: T | None, discovered: T | None) -> T | None:
    if existing is not None and discovered is not None and existing != discovered:
        raise AdapterError(f"conflicting American Express {field}")
    return discovered if discovered is not None else existing


def _consistent_account_reference(existing: str | None, discovered: str | None) -> str | None:
    if existing is None:
        return discovered
    if discovered is None:
        return existing
    existing_suffix = _account_reference_suffix(existing)
    discovered_suffix = _account_reference_suffix(discovered)
    if existing.strip().casefold() == discovered.strip().casefold():
        return discovered
    if existing_suffix is not None and discovered_suffix is not None:
        if existing_suffix == discovered_suffix:
            return discovered
        shorter, longer = sorted((existing_suffix, discovered_suffix), key=len)
        if len(shorter) == 4 and longer.endswith(shorter):
            return discovered if len(discovered_suffix) > len(existing_suffix) else existing
    raise AdapterError("conflicting American Express account reference")


def _account_reference_suffix(value: str) -> str | None:
    match = re.search(r"(\d{4,6})\s*$", value.strip())
    return match.group(1) if match is not None else None


def _cell(row: list[object], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _resolve_description_column(
    headers: list[str],
    *,
    rows: list[list[object]],
    date_col: int,
) -> int:
    """Accept duplicate Amex description aliases only when their values agree."""

    accepted = frozenset(_DESCRIPTION_HEADERS)
    matches = [(index, header) for index, header in enumerate(headers) if header in accepted]
    if not matches:
        raise AdapterError(
            f"required description column missing (expected one of {list(_DESCRIPTION_HEADERS)})"
        )
    if len(matches) == 1:
        return matches[0][0]

    for row in rows:
        if _cell(row, date_col) in (None, ""):
            continue
        values = {
            " ".join(str(_cell(row, index) or "").split()).casefold()
            for index, _header in matches
        }
        if len(values) > 1:
            names = ", ".join(header for _index, header in matches)
            raise AdapterError(f"conflicting description columns: {names}")

    # Alias order is the explicit preference order; Amex's Description wins
    # when the export also includes an equivalent Merchant column.
    return next(
        index
        for alias in _DESCRIPTION_HEADERS
        for index, header in matches
        if header == alias
    )


_FOREIGN_PATTERNS = (
    re.compile(r"^\s*([A-Z]{3})\s*\$?\s*([\d,]+(?:\.\d+)?)\s*$", re.IGNORECASE),
    re.compile(r"^\s*\$?\s*([\d,]+(?:\.\d+)?)\s*([A-Z]{3})\s*$", re.IGNORECASE),
    re.compile(r"^\s*([A-Z]{2})\$\s*([\d,]+(?:\.\d+)?)\s*$", re.IGNORECASE),
)


def _parse_foreign_spend(value: object) -> dict[str, str] | None:
    if value is None or not str(value).strip():
        return None
    text = str(value).strip()
    for index, pattern in enumerate(_FOREIGN_PATTERNS):
        if not (match := pattern.match(text)):
            continue
        if index == 1:
            amount, currency = match.group(1), match.group(2)
        else:
            currency, amount = match.group(1), match.group(2)
        currency = {"US": "USD", "CA": "CAD"}.get(currency.upper(), currency.upper())
        return {"amount": str(parse_decimal(amount)), "currency": currency}
    raise AdapterError(f"invalid Foreign Spend Amount: {value!r}")
