"""Deterministic alias-based adapter for conventional XLSX statements."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from worker.adapters.base import AdapterError
from worker.adapters.generic_csv import GenericCsvAdapter
from worker.models import AccountKind, ParsedFile, ParseResult


class GenericXlsxAdapter:
    format = "xlsx"
    name = "generic_xlsx_v1"

    def detect(self, file: ParsedFile) -> float:
        if file.extension not in {".xlsx", ".xlsm"}:
            return 0.0
        try:
            candidates = self._candidate_sheets(file)
        except Exception:
            return 0.0
        return 0.88 if len(candidates) == 1 else 0.15

    def parse(
        self,
        file: ParsedFile,
        *,
        account_kind: AccountKind = AccountKind.CREDIT_CARD,
    ) -> ParseResult:
        candidates = self._candidate_sheets(file)
        if not candidates:
            raise AdapterError("XLSX contains no conventional transaction table")
        if len(candidates) > 1:
            raise AdapterError("XLSX contains multiple possible transaction tables")
        return GenericCsvAdapter().parse_tabular_rows(
            candidates[0],
            account_kind=account_kind,
        ).model_copy(update={"adapter": self.name})

    @staticmethod
    def _candidate_sheets(file: ParsedFile) -> list[list[list[object]]]:
        workbook = load_workbook(BytesIO(file.content), data_only=True, read_only=True)
        try:
            candidates: list[list[list[object]]] = []
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                if GenericCsvAdapter._locate_header(rows) is not None:
                    candidates.append(rows)
            return candidates
        finally:
            workbook.close()
