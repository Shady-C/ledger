"""AI-assisted mapping for novel CSV/XLSX column layouts.

Only headers and shape-redacted sample cells cross the provider boundary.  The
model proposes a mapping; deterministic parsing, currency/sign validation, and
reconciliation decide whether the mapping may be cached or used.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections.abc import Sequence
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Literal, Protocol

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, ValidationError, field_validator, model_validator

from worker.adapters.base import (
    AdapterError,
    SlashDateOrder,
    extract_statement_metadata,
    infer_direction,
    infer_slash_date_order,
    metadata_with_row_dates,
    normalize_header,
    parse_date,
    parse_optional_flag,
    statement_currency_evidence,
    statement_period_date_values,
)
from worker.llm.provider import LLMDisabledError, LLMProvider, LLMResponseError
from worker.models import (
    AccountKind,
    Direction,
    ParsedFile,
    ParsedTransaction,
    ParseResult,
    ParseStatus,
)
from worker.money import normalize_money
from worker.reconcile import reconcile_statement

DateOrder = Literal["ymd", "mdy", "dmy"]
DecimalSeparator = Literal["dot", "comma"]
AmountSemantics = Literal["credit_card_positive_charges", "asset_positive_inflows", "debit_credit"]


class ColumnMapping(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    booked_date: str
    posted_date: str | None
    description: str
    amount: str | None
    debit: str | None
    credit: str | None
    currency: str | None
    original_amount: str | None
    original_currency: str | None
    fx_fee: str | None
    is_fx_fee: str | None
    reference: str | None
    date_order: DateOrder
    decimal_separator: DecimalSeparator
    amount_semantics: AmountSemantics
    default_currency: str

    @field_validator(
        "booked_date",
        "description",
        "amount",
        "debit",
        "credit",
        "currency",
        "original_amount",
        "original_currency",
        "fx_fee",
        "is_fx_fee",
        "reference",
        "posted_date",
    )
    @classmethod
    def columns_cannot_be_blank(cls, value: str | None) -> str | None:
        if value is None:
            return None
        normalized = value.strip()
        if not normalized:
            raise ValueError("mapped column cannot be blank")
        return normalized

    @field_validator("default_currency")
    @classmethod
    def currency_is_iso_style(cls, value: str) -> str:
        code = value.strip().upper()
        if len(code) != 3 or not code.isalpha():
            raise ValueError("default_currency must be a three-letter code")
        return code

    @model_validator(mode="after")
    def amount_representation_is_exclusive(self) -> ColumnMapping:
        has_amount = self.amount is not None
        has_split = self.debit is not None or self.credit is not None
        if has_amount == has_split:
            raise ValueError("mapping must use either amount or debit/credit columns")
        if self.amount_semantics == "debit_credit" and not has_split:
            raise ValueError("debit_credit semantics require split columns")
        if self.amount_semantics != "debit_credit" and not has_amount:
            raise ValueError("signed amount semantics require one amount column")
        if (self.original_amount is None) != (self.original_currency is None):
            raise ValueError(
                "original_amount and original_currency columns must be mapped together"
            )
        return self


class AdapterMappingStore(Protocol):
    def load_adapter_mapping(
        self, *, account_id: str, format: str, fingerprint: str
    ) -> dict[str, object] | None: ...

    def save_adapter_mapping(
        self,
        *,
        account_id: str,
        format: str,
        fingerprint: str,
        mapping: dict[str, object],
    ) -> None: ...


_MAPPING_SCHEMA: dict[str, object] = {
    "type": "object",
    "properties": {
        "booked_date": {"type": "string"},
        "posted_date": {"type": ["string", "null"]},
        "description": {"type": "string"},
        "amount": {"type": ["string", "null"]},
        "debit": {"type": ["string", "null"]},
        "credit": {"type": ["string", "null"]},
        "currency": {"type": ["string", "null"]},
        "original_amount": {"type": ["string", "null"]},
        "original_currency": {"type": ["string", "null"]},
        "fx_fee": {"type": ["string", "null"]},
        "is_fx_fee": {"type": ["string", "null"]},
        "reference": {"type": ["string", "null"]},
        "date_order": {"type": "string", "enum": ["ymd", "mdy", "dmy"]},
        "decimal_separator": {"type": "string", "enum": ["dot", "comma"]},
        "amount_semantics": {
            "type": "string",
            "enum": [
                "credit_card_positive_charges",
                "asset_positive_inflows",
                "debit_credit",
            ],
        },
        "default_currency": {"type": "string"},
    },
    "required": [
        "booked_date",
        "posted_date",
        "description",
        "amount",
        "debit",
        "credit",
        "currency",
        "original_amount",
        "original_currency",
        "fx_fee",
        "is_fx_fee",
        "reference",
        "date_order",
        "decimal_separator",
        "amount_semantics",
        "default_currency",
    ],
    "additionalProperties": False,
}


class AIColumnMappingService:
    def __init__(self, *, provider: LLMProvider, store: AdapterMappingStore) -> None:
        self.provider = provider
        self.store = store

    def parse(
        self,
        file: ParsedFile,
        *,
        account_id: str,
        account_kind: AccountKind,
        native_currency: str,
    ) -> ParseResult:
        if file.extension not in {".csv", ".xlsx"}:
            return _needs_ai(file, "only novel CSV/XLSX layouts support AI column mapping")
        try:
            rows = _tabular_rows(file)
            header_index = _candidate_header_index(rows)
            headers = [str(value or "").strip() for value in rows[header_index]]
            _validate_headers(headers)
            fingerprint = _fingerprint(
                file.extension[1:],
                headers,
                account_kind=account_kind,
                native_currency=native_currency,
            )
            raw_mapping = self.store.load_adapter_mapping(
                account_id=account_id,
                format=file.extension[1:],
                fingerprint=fingerprint,
            )
            is_new = raw_mapping is None
            if raw_mapping is None:
                raw_mapping = self._propose(
                    headers=headers,
                    sample_rows=rows[header_index + 1 : header_index + 6],
                    account_kind=account_kind,
                    native_currency=native_currency,
                )
            # Mappings cached before Phase 2 remain valid and simply lack
            # optional three-layer/fee columns.
            raw_mapping = {
                "original_amount": None,
                "original_currency": None,
                "fx_fee": None,
                "is_fx_fee": None,
                **raw_mapping,
            }
            mapping = ColumnMapping.model_validate(raw_mapping)
            mapping = _derive_deterministic_formats(
                rows,
                header_index=header_index,
                headers=headers,
                mapping=mapping,
                account_kind=account_kind,
            )
            result = _parse_mapping(
                rows,
                header_index=header_index,
                headers=headers,
                mapping=mapping,
                account_kind=account_kind,
                native_currency=native_currency,
                adapter=f"ai_mapped_{file.extension[1:]}",
            )
            reconciliation = reconcile_statement(result.statement, result.rows)
            if reconciliation.status == "mismatch":
                raise AdapterError("proposed mapping does not reconcile")
            if is_new:
                self.store.save_adapter_mapping(
                    account_id=account_id,
                    format=file.extension[1:],
                    fingerprint=fingerprint,
                    mapping=mapping.model_dump(mode="json"),
                )
            return result
        except (AdapterError, LLMDisabledError, LLMResponseError, ValidationError, ValueError):
            return _needs_ai(file, "AI column mapping could not be validated")

    def _propose(
        self,
        *,
        headers: Sequence[str],
        sample_rows: Sequence[Sequence[object]],
        account_kind: AccountKind,
        native_currency: str,
    ) -> dict[str, object]:
        payload = {
            "headers": list(headers),
            "sample_rows": [
                [_redact_cell(value) for value in row[: len(headers)]] for row in sample_rows
            ],
            "account_kind": account_kind.value,
            "account_currency": native_currency,
        }
        return self.provider.complete(
            system=(
                "Map transaction columns to the supplied schema using exact header strings. "
                "Never calculate amounts or invent columns. A null column means absent. "
                "Date order, decimal convention, sign semantics, and account currency are "
                "validated and replaced deterministically by local code."
            ),
            messages=[{"role": "user", "content": json.dumps(payload, separators=(",", ":"))}],
            schema=_MAPPING_SCHEMA,
            model_tier="capable",
        )


def _derive_deterministic_formats(
    rows: Sequence[Sequence[object]],
    *,
    header_index: int,
    headers: Sequence[str],
    mapping: ColumnMapping,
    account_kind: AccountKind,
) -> ColumnMapping:
    """Resolve every financial parsing semantic from local raw cells only."""

    columns = {
        field: _mapped_index(headers, value)
        for field, value in {
            "booked": mapping.booked_date,
            "posted": mapping.posted_date,
            "amount": mapping.amount,
            "debit": mapping.debit,
            "credit": mapping.credit,
        }.items()
    }
    expected_semantics: AmountSemantics
    if mapping.amount is None:
        expected_semantics = "debit_credit"
    elif account_kind is AccountKind.CREDIT_CARD:
        expected_semantics = "credit_card_positive_charges"
    else:
        expected_semantics = "asset_positive_inflows"

    date_values: list[object] = list(statement_period_date_values(rows[: header_index + 1]))
    amount_values: list[object] = []
    for source in rows[header_index + 1 :]:
        booked = _cell(source, columns["booked"])
        if booked is None or not str(booked).strip():
            continue
        date_values.append(booked)
        posted = _cell(source, columns["posted"])
        if posted is not None and str(posted).strip():
            date_values.append(posted)
        if mapping.amount is not None:
            amount = _cell(source, columns["amount"])
            if amount is not None and str(amount).strip():
                amount_values.append(amount)
        else:
            for field in ("debit", "credit"):
                amount = _cell(source, columns[field])
                if amount is not None and str(amount).strip():
                    amount_values.append(amount)

    date_order = _infer_date_order(date_values)
    decimal_separator = _infer_decimal_separator(amount_values)
    return mapping.model_copy(
        update={
            "date_order": date_order,
            "decimal_separator": decimal_separator,
            "amount_semantics": expected_semantics,
        }
    )


def _infer_date_order(values: Sequence[object]) -> DateOrder:
    slash_order = infer_slash_date_order(values)
    return slash_order or "ymd"


def _infer_decimal_separator(values: Sequence[object]) -> DecimalSeparator:
    if not values:
        raise AdapterError("mapped table contains no amount evidence")
    compatible: set[DecimalSeparator] = {"dot", "comma"}
    for value in values:
        compatible.intersection_update(_decimal_separator_candidates(value))
    if not compatible:
        raise AdapterError("mapped amounts use conflicting or invalid numeric conventions")
    # Plain integers and native numeric XLSX cells are representation-invariant;
    # dot is a harmless canonical marker and every future file is re-inferred.
    return "comma" if compatible == {"comma"} else "dot"


def _decimal_separator_candidates(value: object) -> set[DecimalSeparator]:
    if isinstance(value, (Decimal, int, float)) and not isinstance(value, bool):
        return {"dot", "comma"}
    text = str(value).strip()
    if text.startswith("(") and text.endswith(")"):
        text = text[1:-1]
    text = re.sub(r"(?i)\b[A-Z]{3}\b", "", text)
    text = re.sub(r"[\s$€£¥]", "", text)
    if text[:1] in {"+", "-"}:
        text = text[1:]
    if not text or re.search(r"[^0-9,.]", text):
        return set()
    if "." not in text and "," not in text:
        return {"dot", "comma"} if text.isdigit() else set()

    if "." in text and "," in text:
        decimal_mark = "." if text.rfind(".") > text.rfind(",") else ","
        grouping_mark = "," if decimal_mark == "." else "."
        whole, fraction = text.rsplit(decimal_mark, 1)
        if not 1 <= len(fraction) <= 2 or not fraction.isdigit():
            return set()
        groups = whole.split(grouping_mark)
        if (
            len(groups) < 2
            or not 1 <= len(groups[0]) <= 3
            or not groups[0].isdigit()
            or any(len(group) != 3 or not group.isdigit() for group in groups[1:])
        ):
            return set()
        return {"dot" if decimal_mark == "." else "comma"}

    mark: DecimalSeparator = "dot" if "." in text else "comma"
    symbol = "." if mark == "dot" else ","
    groups = text.split(symbol)
    if any(not group.isdigit() for group in groups):
        return set()
    if len(groups) == 2 and 1 <= len(groups[1]) <= 2:
        return {mark}
    if (
        1 <= len(groups[0]) <= 3
        and len(groups) >= 2
        and all(len(group) == 3 for group in groups[1:])
    ):
        return {"comma" if mark == "dot" else "dot"}
    return set()


def _parse_mapping(
    rows: Sequence[Sequence[object]],
    *,
    header_index: int,
    headers: Sequence[str],
    mapping: ColumnMapping,
    account_kind: AccountKind,
    native_currency: str,
    adapter: str,
) -> ParseResult:
    if mapping.default_currency != native_currency:
        raise AdapterError("proposed mapping currency differs from the selected account")
    if (
        mapping.amount_semantics == "credit_card_positive_charges"
        and account_kind is not AccountKind.CREDIT_CARD
    ):
        raise AdapterError("credit-card sign semantics cannot be used for an asset account")
    if mapping.amount_semantics == "asset_positive_inflows" and not account_kind.is_asset:
        raise AdapterError("asset sign semantics cannot be used for a credit-card account")

    columns = {
        field: _mapped_index(headers, value)
        for field, value in {
            "booked": mapping.booked_date,
            "posted": mapping.posted_date,
            "description": mapping.description,
            "amount": mapping.amount,
            "debit": mapping.debit,
            "credit": mapping.credit,
            "currency": mapping.currency,
            "original_amount": mapping.original_amount,
            "original_currency": mapping.original_currency,
            "fx_fee": mapping.fx_fee,
            "is_fx_fee": mapping.is_fx_fee,
            "reference": mapping.reference,
        }.items()
    }
    slash_order: SlashDateOrder | None = None
    if mapping.date_order == "mdy":
        slash_order = "mdy"
    elif mapping.date_order == "dmy":
        slash_order = "dmy"
    parsed: list[ParsedTransaction] = []
    for source in rows[header_index + 1 :]:
        if not any(value is not None and str(value).strip() for value in source):
            continue
        booked_value = _cell(source, columns["booked"])
        if booked_value in {None, ""}:
            continue
        description = str(_cell(source, columns["description"]) or "").strip()
        if not description:
            raise AdapterError("mapped transaction description is blank")
        amount, direction = _mapped_amount(
            source,
            columns=columns,
            mapping=mapping,
            account_kind=account_kind,
        )
        currency_value = _cell(source, columns["currency"])
        currency = str(currency_value or mapping.default_currency).strip().upper()
        if currency != native_currency:
            raise AdapterError("transaction currency differs from the selected account")
        original_amount_value = _cell(source, columns["original_amount"])
        original_currency_value = _cell(source, columns["original_currency"])
        original_present = original_amount_value is not None and bool(
            str(original_amount_value).strip()
        )
        original_currency_present = original_currency_value is not None and bool(
            str(original_currency_value).strip()
        )
        if original_present != original_currency_present:
            raise AdapterError(
                "mapped original amount and currency must both be present on a row"
            )
        original_amount = None
        original_currency = None
        if original_present:
            original_magnitude = abs(
                _localized_decimal(
                    original_amount_value,
                    separator=mapping.decimal_separator,
                )
            )
            original_amount = -original_magnitude if amount < 0 else original_magnitude
            original_currency = str(original_currency_value).strip().upper()
        resolved_direction = direction or infer_direction(description, amount)
        normalized_description = normalize_header(description)
        has_fx_fee_description = any(
            token in normalized_description
            for token in (
                "foreign exchange fee",
                "fx fee",
                "currency conversion fee",
                "fx commission",
                "exchange commission",
            )
        )
        fx_fee_value = _cell(source, columns["fx_fee"])
        has_inline_fee_value = fx_fee_value is not None and bool(str(fx_fee_value).strip())
        parsed_fx_fee = (
            abs(_localized_decimal(fx_fee_value, separator=mapping.decimal_separator))
            if has_inline_fee_value
            else None
        )
        explicit_standalone = parse_optional_flag(
            _cell(source, columns["is_fx_fee"]), field="standalone FX fee"
        )
        is_fx_fee = (
            explicit_standalone
            if explicit_standalone is not None
            else (
                has_fx_fee_description
                and not original_present
                and (parsed_fx_fee is None or parsed_fx_fee == abs(amount))
            )
        )
        if is_fx_fee and original_present:
            raise AdapterError("a standalone FX-fee row cannot contain original spend")
        if is_fx_fee:
            resolved_direction = Direction.FEE
        fx_fee = parsed_fx_fee if not is_fx_fee else None
        posted_value = _cell(source, columns["posted"])
        reference_value = _cell(source, columns["reference"])
        parsed.append(
            ParsedTransaction(
                booked_date=parse_date(booked_value, slash_order=slash_order),
                posted_date=(
                    parse_date(posted_value, slash_order=slash_order) if posted_value else None
                ),
                description_raw=description,
                amount_native=amount,
                currency_native=currency,
                original_amount=original_amount,
                original_currency=original_currency,
                fx_fee_amount_native=fx_fee,
                is_fx_fee=is_fx_fee,
                external_ref=(str(reference_value).strip() if reference_value else None),
                direction=resolved_direction,
            )
        )
    if not parsed:
        raise AdapterError("mapped table contains no transactions")
    preamble = rows[: header_index + 1]
    metadata = extract_statement_metadata(preamble, slash_date_order=slash_order)
    labelled_statement_currency = statement_currency_evidence(preamble)
    if labelled_statement_currency is not None and labelled_statement_currency != native_currency:
        raise AdapterError("statement balance currency differs from the selected account")
    metadata = metadata.model_copy(update={"currency": native_currency})
    metadata = metadata_with_row_dates(metadata, [row.booked_date for row in parsed])
    return ParseResult(adapter=adapter, rows=tuple(parsed), statement=metadata)


def _mapped_amount(
    row: Sequence[object],
    *,
    columns: dict[str, int | None],
    mapping: ColumnMapping,
    account_kind: AccountKind,
) -> tuple[Decimal, Direction | None]:
    if mapping.amount is not None:
        amount = _localized_decimal(
            _cell(row, columns["amount"]), separator=mapping.decimal_separator
        )
        if mapping.amount_semantics == "asset_positive_inflows":
            return amount, Direction.CREDIT if amount > 0 else Direction.DEBIT
        return amount, None
    debit_value = _cell(row, columns["debit"])
    credit_value = _cell(row, columns["credit"])
    debit_present = debit_value is not None and str(debit_value).strip() != ""
    credit_present = credit_value is not None and str(credit_value).strip() != ""
    if debit_present == credit_present:
        raise AdapterError("mapped row must contain exactly one debit or credit amount")
    magnitude = _localized_decimal(
        debit_value if debit_present else credit_value,
        separator=mapping.decimal_separator,
    )
    if magnitude < 0:
        raise AdapterError("mapped debit/credit magnitudes must be unsigned")
    if account_kind is AccountKind.CREDIT_CARD:
        return (magnitude if debit_present else -magnitude), None
    if debit_present:
        return -magnitude, Direction.DEBIT
    return magnitude, Direction.CREDIT


def _localized_decimal(value: object, *, separator: DecimalSeparator) -> Decimal:
    if value is None or not str(value).strip():
        raise AdapterError("mapped amount is blank")
    text = str(value).strip().replace("(", "-").replace(")", "")
    text = re.sub(r"(?i)\b[A-Z]{3}\b", "", text)
    text = re.sub(r"[^0-9,.-]", "", text)
    if separator == "comma":
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return normalize_money(Decimal(text), field="mapped transaction amount")
    except (InvalidOperation, ValueError) as exc:
        raise AdapterError(f"invalid mapped amount: {value!r}") from exc


def _tabular_rows(file: ParsedFile) -> list[list[object]]:
    if file.extension == ".csv":
        text = file.content.decode("utf-8-sig")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [list(row) for row in csv.reader(StringIO(text), dialect)]
    workbook = load_workbook(BytesIO(file.content), data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _candidate_header_index(rows: Sequence[Sequence[object]]) -> int:
    candidates = [
        (sum(value is not None and bool(str(value).strip()) for value in row), -index, index)
        for index, row in enumerate(rows[:20])
    ]
    if not candidates or max(candidates)[0] < 2:
        raise AdapterError("could not identify a tabular header")
    return max(candidates)[2]


def _validate_headers(headers: Sequence[str]) -> None:
    normalized = [normalize_header(value) for value in headers if value.strip()]
    if len(normalized) < 2 or len(set(normalized)) != len(normalized):
        raise AdapterError("table headers are blank or ambiguous")


def _mapped_index(headers: Sequence[str], value: str | None) -> int | None:
    if value is None:
        return None
    target = normalize_header(value)
    matches = [index for index, header in enumerate(headers) if normalize_header(header) == target]
    if len(matches) != 1:
        raise AdapterError(f"mapped column is missing or ambiguous: {value!r}")
    return matches[0]


def _cell(row: Sequence[object], index: int | None) -> object | None:
    return row[index] if index is not None and index < len(row) else None


def _fingerprint(
    format: str,
    headers: Sequence[str],
    *,
    account_kind: AccountKind,
    native_currency: str,
) -> str:
    encoded = json.dumps(
        {
            "format": format,
            "headers": [normalize_header(value) for value in headers],
            "account_kind": account_kind.value,
            "native_currency": native_currency,
        },
        sort_keys=True,
        separators=(",", ":"),
    ).encode()
    return hashlib.sha256(encoded).hexdigest()


def _redact_cell(value: object) -> object:
    if value is None or not str(value).strip():
        return None
    text = str(value).strip()
    if re.fullmatch(r"[A-Za-z]{3}", text):
        return text.upper()
    if re.fullmatch(r"[()$€£¥+\-\d.,\s]+", text):
        decimal_digits = len(re.sub(r"\D", "", re.split(r"[.,]", text)[-1]))
        sign = "negative" if "-" in text or text.startswith("(") else "positive"
        return f"<number:{sign}:{decimal_digits} decimals>"
    if re.fullmatch(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}", text):
        return re.sub(r"\d", "#", text)
    return "<text>"


def _needs_ai(file: ParsedFile, reason: str) -> ParseResult:
    return ParseResult(
        adapter=f"ai_mapped_{file.extension.lstrip('.') or 'tabular'}",
        status=ParseStatus.NEEDS_AI,
        reason=reason,
    )
